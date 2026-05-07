#!/usr/bin/env python3
"""
detect_format.py — India Credit Lens
--------------------------------------
Pre-flight format check for an incoming SIBC xlsx file.

Reads the raw Excel directly (bypasses parser.py) to detect:
  1. Column layout: outstanding columns vs variation columns
  2. Format classification (supported / unsupported / unknown)
  3. Date ranges in each statement, vs what the pipeline already holds
  4. Series codes: expected vs found, new codes, renamed sectors

Writes format_report.json to the period directory.
Interactive prompt: confirm date range + gaps, then proceed or abort.

Usage:
    python3 detect_format.py rbi-analytics/SIBCS30042026.xlsx
    python3 detect_format.py ... --check     # non-interactive (for run_evals.py)
    python3 detect_format.py ... --dry-run   # detect + print, no write, no prompt

Exit codes:
    0  known format, confirmed by user (interactive) or report already confirmed (--check)
    1  unsupported format, user aborted, or report missing/unconfirmed (--check)
"""

import argparse
import json
import sys
from datetime import date as _date, datetime
from pathlib import Path

import openpyxl

# ── Repo paths ─────────────────────────────────────────────────────────────────

REPO_ROOT     = Path(__file__).resolve().parent.parent
ANALYSIS      = REPO_ROOT / "analysis"
RBI_ANALYTICS = REPO_ROOT / "rbi-analytics"

sys.path.insert(0, str(RBI_ANALYTICS))

try:
    from parser import parse_filename_date, parse_rbi_date, detect_code_and_level
except ImportError as e:
    print(f"ERROR: Cannot import parser.py: {e}", file=sys.stderr)
    sys.exit(1)

# ── ANSI colours ───────────────────────────────────────────────────────────────

GREEN  = "\033[32m"
RED    = "\033[31m"
YELLOW = "\033[33m"
CYAN   = "\033[36m"
BOLD   = "\033[1m"
RESET  = "\033[0m"

def _c(text, colour): return f"{colour}{text}{RESET}"


# ── Known formats ──────────────────────────────────────────────────────────────
#
# Classified by the Row 2 span-header text and outstanding column count.

KNOWN_FORMATS = {
    "monthly_absolute": {
        "description": "Monthly release — 5 absolute outstanding cols + YoY + FY variation",
        "n_outstanding": 5,
        "supported": True,
        "row2_signal": ("year-on-year", "financial year"),   # both substrings must appear
    },
    "annual_3col": {
        "description": "Annual release — 3 absolute outstanding cols + period-over-period variation",
        "n_outstanding": 3,
        "supported": True,
        "row2_signal": (),    # no specific substrings; matches by exclusion
    },
}

# Expected Statement 1 codes, matched against SECTION_DEFS in extract_sibc.py.
# Tuple: (code, display_label, is_priority_sector_memo)
EXPECTED_S1 = [
    ("I",    "Bank Credit",                       False),
    ("II",   "Food Credit",                       False),
    ("III",  "Non-food Credit",                   False),
    ("1",    "Agriculture",                       False),
    ("2",    "Industry",                          False),
    ("2.1",  "Micro and Small",                   False),
    ("2.2",  "Medium",                            False),
    ("2.3",  "Large",                             False),
    ("3",    "Services",                          False),
    ("3.1",  "Transport Operators",               False),
    ("3.2",  "Computer Software",                 False),
    ("3.3",  "Tourism, Hotels & Restaurants",     False),
    ("3.4",  "Shipping",                          False),
    ("3.5",  "Aviation",                          False),
    ("3.6",  "Professional Services",             False),
    ("3.7",  "Trade",                             False),
    ("3.8",  "Commercial Real Estate",            False),
    ("3.9",  "NBFCs",                             False),
    ("3.10", "Other Services",                    False),
    ("4",    "Personal Loans",                    False),
    ("4.1",  "Consumer Durables",                 False),
    ("4.2",  "Housing",                           False),
    ("4.3",  "Advances vs Fixed Deposits",        False),
    ("4.4",  "Advances vs Shares/Bonds",          False),
    ("4.5",  "Credit Card Outstanding",           False),
    ("4.6",  "Education",                         False),
    ("4.7",  "Vehicle Loans",                     False),
    ("4.8",  "Gold Loans",                        False),
    ("4.9",  "Other Personal Loans",              False),
    ("i",    "Agriculture (PSL)",                 True),
    ("ii",   "Micro and Small Enterprises (PSL)", True),
    ("iii",  "Medium Enterprises (PSL)",          True),
    ("iv",   "Housing (PSL)",                     True),
    ("v",    "Educational Loans (PSL)",           True),
    ("vi",   "Renewable Energy (PSL)",            True),
    ("vii",  "Social Infrastructure (PSL)",       True),
    ("viii", "Export Credit (PSL)",               True),
    ("ix",   "Others (PSL)",                      True),
    ("x",    "Weaker Sections (PSL)",             True),
]

# Statement 2 uses all level-2 rows dynamically, so we just flag NEW ones.
# Reference list from the Feb 2026 period.
KNOWN_S2_LEVEL2_CODES = {
    "2.1", "2.2", "2.3", "2.4", "2.5", "2.6", "2.7", "2.8",
    "2.9", "2.10", "2.11", "2.12", "2.13", "2.14", "2.15",
    "2.16", "2.17", "2.18", "2.19",
}


# ── Raw sheet reading ──────────────────────────────────────────────────────────

def read_sheet_structure(ws_data: list) -> dict:
    """
    Extract column headers and data rows directly from openpyxl ws_data.

    Returns:
      outstanding_labels : list of date strings without "/"
      variation_labels   : list of date-range strings with "/"
      variation_type_row : row 2 text (span header), used for format classification
      data_rows          : list of (raw_name, all_numeric_values)
    """
    if len(ws_data) < 6:
        return {"outstanding_labels": [], "variation_labels": [],
                "variation_type_row": "", "data_rows": []}

    # Row 2 (0-indexed): span headers like "Outstanding as on" | "Variation (Year-on-Year)"
    row2_text = " ".join(str(c) for c in ws_data[2] if c).lower()

    # Row 3: individual date / ratio labels
    header_labels = [
        str(c).strip()
        for c in ws_data[3][1:]
        if c is not None and str(c).strip() and str(c).strip().lower() != "none"
    ]

    outstanding = [d for d in header_labels if "/" not in d]
    variation   = [d for d in header_labels if "/" in d]

    # Data rows start at row index 5 (skip title, units, row2, row3/dates, row4/%)
    data_rows = []
    in_priority = False
    for raw_row in ws_data[5:]:
        first = raw_row[0]
        if first is None:
            continue
        name = str(first).strip()
        if not name or name.lower() == "none":
            continue
        name_lower = name.lower()
        if name_lower.startswith("note") or name_lower.startswith("1 ") or name_lower.startswith("(1)"):
            break
        if "priority sector" in name_lower and "(memo)" in name_lower:
            in_priority = True
            continue
        numeric = [raw_row[i + 1] for i in range(len(outstanding) + len(variation))]
        data_rows.append((name, numeric, in_priority))

    return {
        "outstanding_labels": outstanding,
        "variation_labels":   variation,
        "variation_type_row": row2_text,
        "data_rows":          data_rows,
    }


def classify_format(stmt1_info: dict) -> tuple[str, dict]:
    """Return (format_id, format_dict) based on sheet structure."""
    row2 = stmt1_info["variation_type_row"]
    n    = len(stmt1_info["outstanding_labels"])

    for fmt_id, fmt in KNOWN_FORMATS.items():
        signals = fmt["row2_signal"]
        if signals and all(s in row2 for s in signals):
            return fmt_id, fmt
        if not signals and n == fmt["n_outstanding"]:
            return fmt_id, fmt

    return "unknown", {
        "description": f"Unrecognised format ({n} outstanding columns, row2: {row2[:60]})",
        "supported": False,
    }


# ── Date analysis ──────────────────────────────────────────────────────────────

def parse_outstanding_dates(labels: list[str]) -> list[dict]:
    """Parse outstanding date labels into ISO + display strings."""
    result = []
    for lbl in labels:
        dt = parse_rbi_date(lbl)
        result.append({
            "raw":     lbl,
            "iso":     dt.strftime("%Y-%m-%d") if dt else None,
            "display": dt.strftime("%b %Y")    if dt else lbl,
        })
    return result


def load_existing_dates() -> list[str]:
    """Return display dates already in the merged sections (e.g. 'Jan 2024')."""
    merged_path = ANALYSIS / "rbi_sibc" / "merged" / "sections_merged.json"
    if not merged_path.exists():
        return []
    try:
        with open(merged_path) as f:
            data = json.load(f)
        # Use bankCredit as representative section
        for sec in data.get("sections", []):
            if sec["id"] == "bankCredit":
                return [row["date"] for row in sec.get("absoluteData", [])]
    except Exception:
        pass
    return []


# ── Series code inventory ──────────────────────────────────────────────────────

def inventory_codes(data_rows: list) -> dict[str, str]:
    """Return {code: parsed_sector_name} for all rows with a recognised code."""
    found = {}
    for name, _, _ in data_rows:
        code, clean, level = detect_code_and_level(name)
        if code:
            found[code] = clean
    return found


def check_series(s1_found: dict, s2_found: dict) -> dict:
    """Compare found codes against expected codes."""
    expected_s1_codes = {e[0]: e[1] for e in EXPECTED_S1}

    missing = [
        {"code": code, "expected_label": label}
        for code, label in expected_s1_codes.items()
        if code not in s1_found
    ]
    new_s2 = [
        {"code": code, "name": name}
        for code, name in s2_found.items()
        if code not in KNOWN_S2_LEVEL2_CODES
        # Only flag level-2 codes (e.g. "2.20"), not deeper sub-items
        and len(code.split(".")) == 2
    ]
    # Name changes on parent codes are cosmetic — lookup is by code, not label.
    # Only flag renames on leaf/sub-codes (level ≥ 2, i.e. code contains ".").
    name_changes = [
        {"code": code, "expected": expected_s1_codes[code], "found": s1_found[code]}
        for code in expected_s1_codes
        if code in s1_found
        and s1_found[code] != expected_s1_codes[code]
        and "." in str(code)   # sub-codes only; parent renames are intentionally ignored
    ]

    return {"missing": missing, "new_s2": new_s2, "name_changes": name_changes}


# ── Sample table ───────────────────────────────────────────────────────────────

_SAMPLE_CODES_S1 = ["I", "1", "2", "3", "4", "4.5", "4.8"]

def print_sample_table(data_rows: list, outstanding_labels: list[str], title: str):
    """Print a sample of key series with raw outstanding values."""
    parsed_dates = parse_outstanding_dates(outstanding_labels)
    headers = [d["display"] for d in parsed_dates]

    code_rows = {}
    for name, vals, _ in data_rows:
        code, clean, _ = detect_code_and_level(name)
        if code in _SAMPLE_CODES_S1:
            code_rows[code] = (clean, vals[:len(outstanding_labels)])

    col_w = 20
    hdr = f"  {'Series':<32}" + "".join(f"{h:>{col_w}}" for h in headers)
    print(f"\n  {title}")
    print(f"  {'-' * len(hdr.rstrip())}")
    print(hdr)
    print(f"  {'-' * (32 + col_w * len(headers) + 2)}")
    for code in _SAMPLE_CODES_S1:
        if code not in code_rows:
            continue
        clean, vals = code_rows[code]
        label = f"{code}. {clean}"[:32]
        row   = f"  {label:<32}"
        for v in vals:
            if v is None:
                row += f"{'—':>{col_w}}"
            else:
                try:
                    row += f"{float(v):>{col_w},.2f}"
                except (TypeError, ValueError):
                    row += f"{str(v):>{col_w}}"
        print(row)


# ── Main report printer ────────────────────────────────────────────────────────

def print_report(xlsx_name: str, fmt_id: str, fmt: dict,
                 stmt1_info: dict, stmt2_info: dict,
                 date_analysis: dict, series_check: dict):
    w = 72
    print(f"\n{'═' * w}")
    print(f"  {BOLD}SIBC Format Detection — {xlsx_name}{RESET}")
    print(f"{'═' * w}")

    # Format
    supported = fmt.get("supported", False)
    status    = _c("SUPPORTED", GREEN) if supported else _c("NOT YET SUPPORTED", RED)
    print(f"\n  Format   : {BOLD}{fmt_id}{RESET}  [{status}]")
    print(f"  Desc     : {fmt['description']}")

    # Column layout
    print(f"\n  ── Column layout ───────────────────────────────────────────────")
    for stmt_label, info in [("Statement 1", stmt1_info), ("Statement 2", stmt2_info)]:
        out_dates = [d["display"] for d in info["parsed_outstanding"]]
        var_labels = info["variation_labels"]
        print(f"\n  {stmt_label}")
        print(f"    Outstanding ({len(out_dates)}) : {',  '.join(out_dates)}")
        print(f"    Variation   ({len(var_labels)}) : {',  '.join(var_labels)}")

    # Date range analysis
    print(f"\n  ── Date range vs existing pipeline data ────────────────────────")
    existing = date_analysis["existing_dates"]
    print(f"    Merged series currently ends at : {_c(date_analysis['merged_ends_at'], CYAN)}")
    print()
    for stmt_label, dates in [("Statement 1", date_analysis["stmt1_dates"]),
                               ("Statement 2", date_analysis["stmt2_dates"])]:
        print(f"    {stmt_label}:")
        for d in dates:
            disp = d["display"]
            tag  = _c("NEW", GREEN) if disp not in existing else _c("overlap", YELLOW)
            print(f"      {disp:<14} [{tag}]")

    # Series check
    print(f"\n  ── Series inventory ────────────────────────────────────────────")
    sc = series_check
    if not sc["missing"] and not sc["new_s2"] and not sc["name_changes"]:
        print(f"  {_c('✓  All expected codes present. No new series detected.', GREEN)}")
    else:
        if sc["missing"]:
            print(f"\n  {_c('MISSING expected codes (pipeline will skip these series):', RED)}")
            for m in sc["missing"]:
                print(f"    ✗  {m['code']:<8} {m['expected_label']}")
        if sc["new_s2"]:
            print(f"\n  {_c('NEW Statement 2 codes not in SECTION_DEFS (will be auto-added to industryByType):', YELLOW)}")
            for n in sc["new_s2"]:
                print(f"    +  {n['code']:<8} {n['name']}")
        if sc["name_changes"]:
            print(f"\n  {_c('RENAMED series (code matches but sector name differs):', YELLOW)}")
            for nc in sc["name_changes"]:
                print(f"    ~  {nc['code']:<8} expected: '{nc['expected']}'  found: '{nc['found']}'")

    # Sample values
    print(f"\n  ── Sample outstanding values (raw, as extracted from Excel) ────")
    print_sample_table(stmt1_info["data_rows"], stmt1_info["outstanding_labels"],
                       "Statement 1 — key series")

    print(f"\n{'═' * w}\n")


# ── Interactive prompt ─────────────────────────────────────────────────────────

def prompt_user(fmt_id: str, supported: bool) -> bool:
    """
    Show what will happen next and ask for confirmation.
    Returns True if user confirms (proceed), False to abort.
    """
    if supported:
        print("  The format is SUPPORTED. Extraction can proceed once you confirm.")
    else:
        print(f"  {_c('⚠  Format is NOT YET SUPPORTED.', YELLOW)}")
        print("  Confirming will record this detection and BLOCK extraction until")
        print("  parser support is added for this format.")

    print()
    print("  Please review the date ranges and series inventory above.")
    print()
    print("  [A] Confirm — record this format report (extraction will proceed if supported,")
    print("                block if unsupported)")
    print("  [B] Abort   — do not record; investigate raw file first")
    print()

    try:
        choice = input("  Choice (A/B): ").strip().upper()
    except (EOFError, KeyboardInterrupt):
        print("\n  No terminal — defaulting to [B] abort.")
        return False

    return choice == "A"


# ── format_report.json ────────────────────────────────────────────────────────

def write_report(period_dir: Path, payload: dict):
    period_dir.mkdir(parents=True, exist_ok=True)
    path = period_dir / "format_report.json"
    with open(path, "w") as f:
        json.dump(payload, f, indent=2)
    print(f"\n  ✅ Saved → {path.relative_to(REPO_ROOT)}")


# ── Main ───────────────────────────────────────────────────────────────────────

def detect(xlsx_path: Path, check_mode: bool = False, dry_run: bool = False) -> int:
    """
    Run format detection.
    Returns 0 (OK/confirmed) or 1 (unsupported/aborted/missing).
    """
    # ── Check mode: just read the saved report ──────────────────────────────
    if check_mode:
        report_date = parse_filename_date(xlsx_path)
        if report_date is None:
            print(f"  [detect_format] Cannot parse date from filename: {xlsx_path.name}",
                  file=sys.stderr)
            return 1
        period_dir  = ANALYSIS / "rbi_sibc" / report_date.strftime("%Y-%m-%d")
        report_path = period_dir / "format_report.json"
        if not report_path.exists():
            print(f"  [detect_format] format_report.json missing for {period_dir.name} — "
                  f"run detect_format.py first", file=sys.stderr)
            return 1
        with open(report_path) as f:
            report = json.load(f)
        if not report.get("confirmed_by_user"):
            print(f"  [detect_format] format not confirmed by user", file=sys.stderr)
            return 1
        if not report.get("supported"):
            print(f"  [detect_format] format '{report.get('format_id')}' not supported — "
                  f"add parser support before extraction", file=sys.stderr)
            return 1
        return 0

    # ── Full detection ──────────────────────────────────────────────────────
    if not xlsx_path.exists():
        print(f"ERROR: File not found: {xlsx_path}", file=sys.stderr)
        return 1

    report_date = parse_filename_date(xlsx_path)
    if report_date is None:
        print(f"ERROR: Cannot parse report date from filename: {xlsx_path.name}",
              file=sys.stderr)
        return 1
    data_date  = report_date.strftime("%Y-%m-%d")
    period_dir = ANALYSIS / "rbi_sibc" / data_date

    # Load workbook
    wb = openpyxl.load_workbook(str(xlsx_path), read_only=True, data_only=True)
    raw = {}
    for sheet_name in wb.sheetnames:
        ws_data = [tuple(row) for row in wb[sheet_name].iter_rows(values_only=True)]
        raw[sheet_name] = ws_data
    wb.close()

    stmt1_raw = raw.get("Statement 1", [])
    stmt2_raw = raw.get("Statement 2", [])

    stmt1_struct = read_sheet_structure(stmt1_raw)
    stmt2_struct = read_sheet_structure(stmt2_raw)

    fmt_id, fmt = classify_format(stmt1_struct)

    # Parse dates
    stmt1_parsed = parse_outstanding_dates(stmt1_struct["outstanding_labels"])
    stmt2_parsed = parse_outstanding_dates(stmt2_struct["outstanding_labels"])

    stmt1_info = {**stmt1_struct, "parsed_outstanding": stmt1_parsed}
    stmt2_info = {**stmt2_struct, "parsed_outstanding": stmt2_parsed}

    # Date range vs pipeline
    existing      = load_existing_dates()
    merged_end    = existing[-1] if existing else "—"
    date_analysis = {
        "existing_dates":  existing,
        "merged_ends_at":  merged_end,
        "stmt1_dates":     stmt1_parsed,
        "stmt2_dates":     stmt2_parsed,
    }

    # Series inventory
    s1_codes = inventory_codes(stmt1_struct["data_rows"])
    s2_codes = inventory_codes(stmt2_struct["data_rows"])
    series   = check_series(s1_codes, s2_codes)

    # Print report
    print_report(xlsx_path.name, fmt_id, fmt,
                 stmt1_info, stmt2_info, date_analysis, series)

    if dry_run:
        print("  [dry-run] No report written.\n")
        return 0 if fmt.get("supported") else 1

    # Prompt
    confirmed = prompt_user(fmt_id, fmt.get("supported", False))

    # Build JSON payload
    payload = {
        "xlsx":             xlsx_path.name,
        "data_date":        data_date,
        "period_dir":       str(period_dir.relative_to(REPO_ROOT)),
        "detected_at":      _date.today().isoformat(),
        "format_id":        fmt_id,
        "format_description": fmt.get("description", ""),
        "supported":        fmt.get("supported", False),
        "confirmed_by_user": confirmed,
        "stmt1": {
            "outstanding_labels": stmt1_struct["outstanding_labels"],
            "outstanding_iso":    [d["iso"] for d in stmt1_parsed],
            "variation_labels":   stmt1_struct["variation_labels"],
            "n_outstanding":      len(stmt1_struct["outstanding_labels"]),
        },
        "stmt2": {
            "outstanding_labels": stmt2_struct["outstanding_labels"],
            "outstanding_iso":    [d["iso"] for d in stmt2_parsed],
            "variation_labels":   stmt2_struct["variation_labels"],
            "n_outstanding":      len(stmt2_struct["outstanding_labels"]),
        },
        "date_analysis": {
            "merged_ends_at":  merged_end,
            "stmt1_new_dates": [d["display"] for d in stmt1_parsed
                                if d["display"] not in existing],
            "stmt2_new_dates": [d["display"] for d in stmt2_parsed
                                if d["display"] not in existing],
        },
        "series_check": {
            "stmt1_codes_found": list(s1_codes.keys()),
            "stmt2_codes_found": list(s2_codes.keys()),
            "missing_expected":  series["missing"],
            "new_s2_codes":      series["new_s2"],
            "name_changes":      series["name_changes"],
        },
    }

    write_report(period_dir, payload)

    if not confirmed:
        print("  Aborted — format report saved with confirmed=false.\n")
        return 1

    if not fmt.get("supported"):
        print(f"  Format '{fmt_id}' recorded but UNSUPPORTED.")
        print("  Add parser support, then re-run extract_sibc.py.\n")
        return 1

    return 0


# ── CLI ────────────────────────────────────────────────────────────────────────

if __name__ == "__main__":
    ap = argparse.ArgumentParser(description="Detect SIBC xlsx format before extraction")
    ap.add_argument("xlsx",      help="Path to SIBC .xlsx file")
    ap.add_argument("--check",   action="store_true",
                    help="Non-interactive: read saved report and exit 0/1")
    ap.add_argument("--dry-run", action="store_true",
                    help="Detect and print; do not write report or prompt")
    args = ap.parse_args()

    sys.exit(detect(
        xlsx_path=Path(args.xlsx),
        check_mode=args.check,
        dry_run=args.dry_run,
    ))
